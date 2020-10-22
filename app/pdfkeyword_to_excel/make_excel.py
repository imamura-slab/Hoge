import openpyxl as xl
import csv
import numpy as np


## flags {
TOP3 = False
TOP2 = True
TOP1 = True
## flags }


def main(dargs):
    input_excel_name  = dargs["INPUT_EXCEL_NAME"]
    output_excel_name = dargs["OUTPUT_EXCEL_NAME"]
    input_csv_name    = dargs["INPUT_CSV_NAME"]
    
    
    ## READ Excel file
    wb = xl.load_workbook(input_excel_name)
    # ws = wb.copy_worksheet(wb["Sheet1"])
    # ws.title = "Sheet2"
    ws = wb["Sheet1"]


    # wb = xl.load_workbook(input_excel_name)
    # ws = wb.worksheets[0]
    # wb.copy_worksheet(ws)
    
    
    ## DELETE previous keywords
    keyword = ws[dargs["KEYWORD_RANGE"]]
    for i in range(len(keyword)):
        for j in range(len(keyword[i])):
            keyword[i][j].value=""
            
            
    ## ADD auto filter
    ws.auto_filter.ref = dargs["AUTO_FILTER_RANGE"]

    
    ## COPY color of auto filter
    target = ws["B10:DI10"]
    for i in range(len(target)):
        for j in range(len(target[i])):
            target[i][j]._style = ws["A10"]._style


    ## 学生番号抜き出し
    # target = ws["C10:C89"]
    # for i in range(len(target)):
    #     for j in range(len(target[i])):
    #         print(target[i][j].value)
    # exit()

            
            
    ## READ keyword from csv file and WRITE
    keyword_num = np.array([], dtype="int8")
    row = dargs["KEYWORD_START_ROW_NUM"]
    col = dargs["KEYWORD_START_COLUMN_NUM"]
    with open(input_csv_name) as f:
        reader = csv.reader(f)
        for r, data in enumerate(reader):
            print('input ID:', data[0])
            ### ここで学籍番号 or name チェック
            student_ID = ws.cell(row+r, 3).value
            #print('ID:', student_ID)
            while data[0] != str(student_ID):
                keyword_num = np.append(keyword_num, None)
                ws.cell(row+r, 9).value = ""
                row         = row + 1
                student_ID  = ws.cell(row+r, 3).value
                if student_ID is None:
                    print('ERROR : Not match ID')
                    exit()
                        
                        
            keyword_num = np.append(keyword_num, len(data))
            for index in range(len(data)-1):
                # ws.cell(row=row+r, column=col+index).value = data[index]
                ws.cell(row+r, col+index).value = data[index+1]
    

    # keyword_num = np.append(keyword_num, None)                
    # print(keyword_num)
    # print(keyword_num[3])


    ## SORT #keywords and select TOP3 score
    # REMOVE None
    filtered_list = [e for e in keyword_num if e is not None]
    if len(filtered_list) < 3:
        print("can\'t select TOP3. too few kinds of #keywords")
        first  = 10000
        second = 10000
        third  = 10000
    else:
        # SORT
        first  = sorted(set(filtered_list))[-1]
        second = sorted(set(filtered_list))[-2]
        third  = sorted(set(filtered_list))[-3]
        print(first, second, third)

    
    
    ## DRAW TOP3 and NOT submit persons { 
    ##----------------------------------##
    ##       TOP3 -> fill Yellow        ##
    ## NOT submit -> red                ##
    ##----------------------------------##
    yellow_fill = xl.styles.PatternFill(patternType='solid', fgColor='ffff00', bgColor='d7d7d7')
    red_char   = xl.styles.fonts.Font(color='FF0000')
    row_offset = dargs["KEYWORD_START_ROW_NUM"]
    
    for i, score in enumerate(keyword_num):
        if score is None:     # NOT submit
            start = "A" + str(row_offset+i)
            end   = dargs["KEYWORD_END_COLUMN"] + str(row_offset+i)
            cell  = ws[start:end]
            for i in range(len(cell)):
                for j in range(len(cell[i])):
                    cell[i][j].font = red_char
        else:
            if score < third: # NOT TOP3
                pass
            else:
                start = "A" + str(row_offset+i)
                end   = dargs["KEYWORD_END_COLUMN"] + str(row_offset+i)
                cell  = ws[start:end]
                if score < second:   # 3rd
                    for i in range(len(cell)):
                        for j in range(len(cell[i])):
                            cell[i][j].fill = yellow_fill
                elif score < first:  # 2nd
                    for i in range(len(cell)):
                        for j in range(len(cell[i])):
                            cell[i][j].fill = yellow_fill
                else:                # 1st
                    for i in range(len(cell)):
                        for j in range(len(cell[i])):
                            cell[i][j].fill = yellow_fill
    ## DRAW TOP3 and NOT submit persons }
        

    
    ## multi columns to one column {
    row = dargs["KEYWORD_START_ROW_NUM"]
    col = dargs["KEYWORD_START_COLUMN_NUM"]
    count = 0
    
    for i in range(104):
        for j in range(89):
            if ws.cell(row+i, col+j).value == "":
                break
            #ws.cell(100+i*104+j, 3).value = ws.cell(row+i, col+j).value
            ws.cell(100+count, 3).value = ws.cell(row+i, col+j).value
            count += 1
    ## multi columns to one column }
    

    ## COUNT KEYWORD { ## must CHANGE read mode
    # for i in range(count):
    #     target = "=COUNTIF(J11:DI89, C" + str(100+i) + ")"
    #     ws.cell(100+i, 4).value = target
    ## COUNT KEYWORD }

    

    
    
    ## WRITE Excel file                
    wb.save(output_excel_name)

    

if __name__ == '__main__':
    dargs = {
        "INPUT_EXCEL_NAME"         : "template.xlsx",
        "OUTPUT_EXCEL_NAME"        : "output.xlsx",
                                  
        #"INPUT_CSV_NAME"           : "result.csv",
        "INPUT_CSV_NAME"           : "id.csv",
                                  
        # "AUTO_FILTER_RANGE"        : "A10:DI81",
        # "KEYWORD_RANGE"            : "J11:DI81",
        "AUTO_FILTER_RANGE"        : "A10:DI89",
        "KEYWORD_RANGE"            : "J11:DI89",
        
        "KEYWORD_START_ROW_NUM"    : 11,
        "KEYWORD_START_COLUMN_NUM" : 10,

        "KEYWORD_START_COLUMN"     : "J",
        "KEYWORD_END_COLUMN"       : "DI",
        "KEYWORD_START_ROW"        : 11,
        "KEYWORD_END_ROW"          : 89,

        "ONE_COLUMN_RANGE"         : "C100:C9356",
    }
    

    main(dargs)


    

